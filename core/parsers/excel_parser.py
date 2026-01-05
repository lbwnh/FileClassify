"""
Excel 解析器实现。

本模块提供 Excel 文件解析功能。
"""

from .base import BaseParser
from typing import Dict


class ExcelParser(BaseParser):
    """
    Excel 文件解析器（XLSX, XLS, CSV）。

    使用 openpyxl 或 pandas 进行文本提取。
    """
    
    def extract_text(self) -> str:
        """
        从 Excel 文件中提取文本内容。

        返回:
            str: 提取的文本内容
        """
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"[Sheet: {sheet_name}]")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        text_content.append(row_text)
            
            return '\n'.join(text_content)
        
        except ImportError:
            try:
                import pandas as pd
                
                if self.file_path.suffix.lower() == '.csv':
                    df = pd.read_csv(self.file_path)
                    return df.to_string()
                else:
                    excel_file = pd.ExcelFile(self.file_path)
                    text_content = []
                    
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        text_content.append(f"[Sheet: {sheet_name}]")
                        text_content.append(df.to_string())
                    
                    return '\n'.join(text_content)
            
            except ImportError:
                return f"[Excel Parser] Cannot extract text: openpyxl or pandas not installed.\nFile: {self.file_path.name}"
        
        except Exception as e:
            return f"[Excel Parser] Error extracting text: {str(e)}\nFile: {self.file_path.name}"
    
    def extract_metadata(self) -> Dict[str, any]:
        """
        从 Excel 文件中提取元数据。

        返回:
            dict: Excel 元数据
        """
        metadata = {
            'title': None,
            'author': None,
            'subject': None,
            'keywords': None,
            'created_date': None,
            'modified_date': None,
            'sheet_count': 0,
            'sheet_names': []
        }
        
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            props = workbook.properties
            
            metadata['title'] = props.title
            metadata['author'] = props.creator
            metadata['subject'] = props.subject
            metadata['keywords'] = props.keywords
            metadata['created_date'] = props.created
            metadata['modified_date'] = props.modified
            
            metadata['sheet_count'] = len(workbook.sheetnames)
            metadata['sheet_names'] = workbook.sheetnames
        
        except Exception as e:
            metadata['error'] = str(e)
        
        return metadata
